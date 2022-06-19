#pip install pillow
#pip install openpyxl

""" 
RiskyBusiness will scour the relevant downtime excel sheet on the s/Drive, and create an array of 
RISK ratings and a matching array of equipment numbers. Those arrays will be sent to LIMBLE via API (seperate function/file)
 """

# - THIS IS WHAT IS NEXT - ##############
""" 
6.18 -  P067 - Adamatic / SL
        P068 - Stringline   - 26
        P069 - MainLine     - 25
        P070 - 
        P071 - BreadLine    - 23
        P072 - MuffinLine   - 24
        Organize info from downtime.xlsx, col D, into 4 categories (above), 
        then apply RISK to those downtime minutes, then update assets with new RISK info via API



 """
#########################################
from this import d, s
from openpyxl import load_workbook
import http.client
import json
from http.client import HTTPSConnection
from base64 import b64encode
import numpy as np
from openpyxl import load_workbook

def getAssets():
    conn = http.client.HTTPSConnection("api.limblecmms.com", 443)
    payload = ""
    userAndPass = b64encode(
        b"DCHBKYZF5NMXHCV8AG4M1J53DFDONO8Z:WV0KBNNCRLP0SO3CYZMOGFQATYTPG2Y").decode("ascii")

    headers = {
        'Authorization': 'Basic %s' % userAndPass,
    }
    conn.request("GET", "/v2/assets", payload, headers)

    res = conn.getresponse()
    data = res.read()

    json_output = json.loads(data)
    #print(json_output)

    parentID = []
    eqName = []
    eqID = []
    sevArr = []

    for i in json_output:
        #print(i['name'])
        #print(i['assetID'])

        eqName.append(i['name']) #DONT REALL NEED THIS
        eqID.append(i['assetID'])
        parentID.append(i['parentAssetID'])
    #END FOR-i

    with open('assets.json', 'w') as f: #THIS builds a json file called assets.json. It is not necessary.
        json.dump(json_output, f)
    #END WITH OPEN

    #print(eqName)
    #print(eqID)
    #print(parentID)

    ###HERE I NEED TO TRANSLATE PARENTID INTO SEVERITY RATING USING IF/ELIF STATEMENT
    buildSev()
    ##"""
    for j, eq in enumerate(eqID):
        print(eq)
        #severity = sevArr[j]

        payload = json.dumps({
            "severity": severity
        })

        conn.request("PATCH", "/v2/assets/fields/" + str(severity)  + "/?fields=7&assetID=" + str(eq), payload, headers)
        res = conn.getresponse()
        data = res.read()
        #print(data.decode("utf-8"))

    #END FOR-i, eq
    ##"""
    return eqName
#END getAssets()

def buildSev():
    wb = load_workbook(filename="dt2.xlsx")
    ws = wb.active

    sl = 0
    bl = 0
    el = 0
    ml = 0
    #open Downtime.xlsx, loop down col D
    #dCell = value of cell d on current row 
    #gcell = val of cell g on current row
    for i, dCell in enumerate(ws["D"]):
        thisG = "G" + str(i+1)
        gCell = ws[thisG]

        gCell = gCell.value
        dCell = dCell.value

        #print(dCell)
        #print(gCell)
        if dCell == "P068" or dCell == "P067":
            sl = sl + gCell
        elif dCell == "P069":
            ml = ml + gCell
        elif dCell == "P071":
            bl = bl + gCell
        elif dCell == "P072":
            el = el + gCell
        #END IF-dCell
    #END FOR-i, dCell
    sever
    print(bl)
    print(ml)
    print(el)
    print(sl)
#END buildSev()    

getAssets()