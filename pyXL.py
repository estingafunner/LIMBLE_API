#pip install pillow
#pip install openpyxl

# - THIS IS WHAT IS NEXT - ##############
""" 
6.12 I need to make another function (or set of functions) to build the ASSET import sheet for Limble
I don't think there is a sheet that has ALL of the equipment WITH RISK AND SEVERITY.
Should I wait for the equipement before filling the PM import sheet? Is there any adventatge?
6.13 Keep working on matching pmNum to assetID
6.14 Complete!
 """
#########################################

import numpy as np
import re
from openpyxl import load_workbook

def fromPMBooks(): #WORKS!6.13 - This will strip the PM Books for equipment name/number, task info, and frequence
    wb = load_workbook(filename="PM Books.xlsm")
    #ws = wb.active #I don't think I need this. NOPE! DIDN'T NEED IT!
    

    eqArr = []
    taskArr = []
    freqArr = [] #THESE THREE ARRAYS will eventually be stacked into a single array like this: fourKey = np.stack((fcArr, finPrioArr, taskArr), axis=1) TURNS OUT I DON'T NEED TO STACK THE ARRAYS AFTERALL.

    exCounter = 0   #DITCH this, add a statement to first IF... AND len(eqArr) < 100 
                    #The point of this is to meet limble's PM import maximum of 100 tasks
                    #THERE IS no need to limit the number of tasks in this function!
                    #instead, another function will split the array of all tasks into 100-task chunks and export to limble that way

    for ws in wb.worksheets:

        if "BOOK" in ws.title or "Support 1" in ws.title or "Support 2" in ws.title:

            #print(ws.title)
            initialCell = 0
            finCell = 0
            ###PROBABLY change this to a loop through Range(3(wherever FREQ is) to len(ws['c'])
            for index, cell in enumerate(ws['C']):
                
                #print(cell.value)
                if cell.value is None and initialCell == 0:
                    initialCell = index 
                    

                elif cell.value is None and initialCell != 0:
                    initialCell = initialCell + 1
                    finCell = index
                    #print(finCell)

                    eqFind = "B" + str(initialCell)
                    eqRaw = ws[eqFind].value
                    #print(eqRaw)

                    for i in range((initialCell + 1) , (finCell + 1)):
                        taski = "B" + str(i)
                        freqi = "C" + str(i)
                        eqArr.append(eqRaw)
                        taskArr.append(ws[taski].value)
                        freqArr.append(ws[freqi].value)
                    #END FOR-i range

                    #print(eqArr)
                    #print(taskArr)
                    

                    #initialCell - 1 of B is Equipment RAW, ad
                    #from initialCell to finCell, add to array Bi, Ci, and EqRaw(for every i)

                    initialCell = finCell 

                    #Loop from initialCell to finCell over column B and D
                    #build array as [equipment..., task, freq]

                    #ALL NEEDS TO END WHEN ARRAY has 100 entries (due to the import limitation of LIMBLE)

                    #THIS NEEDS TO END with  initialCell = 0 and finCell = 0
                    """ cell.value == "FREQ":
                    initialCell = index
                    print("1 - initialCell - " ) """ # this was made obsolete by checking if initialCell == 0



                #END if-cell "FREQ"

            #END for-cells    

        #END if-"BOOK"-ws.title    

    #END for-ws-wb.worksheets
    
    ##print(len(list(set(eqArr))))
    #print(list(set(eqArr)))
    ##print(len(list(set(taskArr))))
    #print(list(set(taskArr)))
    ##print(len(np.unique(np.sort(freqArr))))
    ##print(np.unique(np.sort(freqArr)))
    #print(eqArr)
    
    pmbArr = np.stack((eqArr, taskArr, freqArr), axis=1) #Decided to skip this and send all 
                    #three arrays to 'buildimportsheet()' individually. Simpler. More Simpler.
    buildImportSheet(eqArr, taskArr, freqArr)

    #END fromPMBooks()

def buildImportSheet(eqArr, taskArr, freqArr): #WORKS! 6.13 - Takes Equipment info, frequency, and task descriptions and funnels them into excel for import into LIMBLE/CMMS
    wb = load_workbook(filename="sample_PM_Templates_list.xlsx")
    wbb = load_workbook(filename="Assets-Export-Limble.xlsx")
    ws = wb.active
    wss = wbb.active

    for x, thing in enumerate(eqArr):
        #print(thing)
        if thing is not None:
            PMTask = taskArr[x]
            PMFreq = freqArr[x]
            
            clean1 = thing.find('#') + 1
            #print(clean1)
            clean2 = thing[clean1:(len(thing))]
            #print(clean2)
            clean3 = re.sub(r"\s{2,}"," ",clean2) #clean2.strip()
            clean4 = clean3.strip()
            #print(clean4)
            #print(" - ")
            eqNum = clean4[:3]
            PMassetID = 0
           
            for v in range(2, len(wss["B"])):

                eqNom = wss[("B" + str(v))].value
                eqNom = eqNom[:3]

                if eqNum == eqNom:
                    assetID = wss[("A" + str(v))].value  
                    print(assetID)
                    print(eqNom)
                    print(eqNum)
                    print(" - ")
                    PMassetID = assetID

                #END IF-LOOPINT 
            #END FOR-V

            PMName = str(eqNum) + "." + str(PMFreq) + "-" + str(PMTask)[:7] + "..." #   "PM." + 
            
            #print(PMName)
            #print(PMTask)
            #print(PMFreq)
            #print(" - ")

            #!!!!!!! This is where code goes to move the three above-printed variables into "PM Import.xlsm"
            lastRow = len(ws['A']) + 1
            ws.cell(row = lastRow, column = 1).value = PMName #Column A
            ws.cell(row = lastRow, column = 3).value = PMTask #Column C
            ws.cell(row = lastRow, column = 13).value = PMassetID #Column M
            ws.cell(row = lastRow, column = 15).value = PMFreq #Column O
        #END IF-THING is NOT NONE

    #END FOR-THING
    wb.save("PMImport4Limble.xlsx")
    #print("I Need this break to prevent Filetype issues for the PM Import.xlsm file.")
    #print(pmbArr)

    #END buildImportSheets()

fromPMBooks()