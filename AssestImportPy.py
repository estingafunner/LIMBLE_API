# - THIS IS WHAT IS NEXT - ##############
""" 
6.13 - Just starting off! Pulling and sorting data from assetsRaw.xlsx into Sample Asset Import.xlsx
 """
#########################################

from operator import mod
from tkinter.font import names
import numpy as np
import re
from openpyxl import load_workbook

def eqNameBuild(): #WORKS! 6.13 - pull info from assetsRaw to build array of equipNum - AssetName, Format Example: "964 - North Elevator"
    namesArr = []
    makeArr = []
    modelArr = []
    serialArr = []
    shortParentArr = [] #didnt use
    parentArr = []

    wb = load_workbook(filename="assetsRaw.xlsx")
    ws = wb["Asset"]

    for i, cell in enumerate(ws["A"]):
        if cell.value != "Asset Name":
            z = i + 1
            eqName = "A" + str(z)
            #print(eqName)
            #print(ws[eqName].value)
            longACode = "B" + str(z)
            longPCode = "H" + str(z) #didnt need

            x = ws[longACode].value.rsplit("-")
            #print(x)

            parent = (x[((len(x) - 2))])
            eqNum = (x[((len(x) - 1))])
            #print(eqNum)

            if len(x) > 3 and eqName != "A602" and eqNum[3:].isalpha():
                eqNum = eqNum[:3]
            elif len(x) > 3 and eqName != "A602" and eqNum[:1].isalpha():
                parent = (x[((len(x) - 3))])
                eqNum = (x[((len(x) - 2))])
                eqNum = eqNum[:3]
            #END EQNUM    

            finName = eqNum + " - " + ws[eqName].value
            #print(finName)

            namesArr.append(finName)
            parentArr.append(parent)
            make = "D" + str(z)
            makeArr.append(ws[make].value)
            model = "E" + str(z)
            modelArr.append(ws[model].value)
            serial = "F" + str(z)
            serialArr.append(ws[serial].value)

            #
        #END IF CELL.VALUE
    #END FOR-Z

    for eqNom in namesArr: #get fpn (full parent name) from parent by matching to namesArr[:3]
        #print(eqNom)
        for k, parr in enumerate(parentArr):
            if eqNom[:(len(parr))] == parr:
                parentArr[k] = (eqNom)
                #print(eqNom)
            #END IF-EQNOM
        #END FOR-K
    #END FOR-J

    #print(parentArr)
    #print(len(namesArr))
    importBuild(namesArr, parentArr, makeArr, modelArr, serialArr)
#END eqNameBuild()

def importBuild(namesArr, parentArr, makeArr, modelArr, serialArr):
    wb = load_workbook(filename="Sample Asset Import.xlsx")
    ws = wb.active

    for i, nom in enumerate(namesArr):
        x = 2 + i

        ws[("A" + (str(x)))].value = namesArr[i]
        ws[("B" + (str(x)))].value = parentArr[i]
        ws[("C" + (str(x)))].value = makeArr[i]
        ws[("D" + (str(x)))].value = modelArr[i]
        ws[("E" + (str(x)))].value = serialArr[i]
        print(" - ")
    #END FOR-I NOM   
    wb.save("Asset Import Pyd.xlsx")
#END importBuild()

eqNameBuild()